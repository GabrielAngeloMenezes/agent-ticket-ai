import os
import re
import json
import base64
import shutil
from datetime import datetime

import fitz
import pytesseract
from PIL import Image, ImageFilter, ImageOps, ImageEnhance
from flask import Flask, render_template, request, redirect, session, send_file
from openai import OpenAI
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from openpyxl import load_workbook

from agent.supabase_client import supabase

if os.name == "nt":
    pytesseract.pytesseract.tesseract_cmd = (
        r"C:\Program Files\Tesseract-OCR\tesseract.exe"
    )

client = OpenAI(
    base_url="https://openrouter.ai/api/v1",
    api_key=os.getenv("OPENROUTER_API_KEY")
)

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "agent-secret-key")

UPLOAD_FOLDER = "uploads"
EXPORT_FOLDER = "exports"
TEMPLATE_FOLDER = "templates_excel"
MODELO_COTACAO = os.path.join(TEMPLATE_FOLDER, "modelo_cotacao.xlsx")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(EXPORT_FOLDER, exist_ok=True)


def usuario_logado():
    return "usuario_nome" in session


def buscar_usuario(nome):
    resposta = supabase.table("usuarios").select("*").eq("nome", nome).execute()
    if resposta.data:
        return resposta.data[0]
    return None


def criar_usuario(nome, senha):
    senha_hash = generate_password_hash(senha)
    supabase.table("usuarios").insert({
        "nome": nome,
        "senha": senha_hash
    }).execute()


def buscar_historico(nome, limite=20):
    resposta = (
        supabase
        .table("tickets")
        .select("*")
        .eq("user_id", nome)
        .order("id", desc=False)
        .limit(limite)
        .execute()
    )
    return resposta.data


def responder_chat(mensagem, historico):
    mensagens = [
        {
            "role": "system",
            "content": (
                "Voce e um agente de IA amigavel e inteligente. "
                "Converse naturalmente. "
                "Lembre do contexto anterior. "
                "Responda de forma clara e util."
            )
        }
    ]

    for item in historico:
        mensagens.append({"role": "user", "content": item["ticket"]})
        mensagens.append({"role": "assistant", "content": item["categoria"]})

    mensagens.append({"role": "user", "content": mensagem})

    resposta = client.chat.completions.create(
        model="openrouter/free",
        messages=mensagens,
        temperature=0.7,
        max_tokens=300
    )

    conteudo = resposta.choices[0].message.content

    if conteudo is None:
        return "Nao consegui responder agora."

    return conteudo.strip()


def salvar_mensagem(nome, usuario, resposta):
    supabase.table("tickets").insert({
        "user_id": nome,
        "ticket": usuario,
        "categoria": resposta
    }).execute()


def limpar_codigo(codigo):
    codigo = str(codigo).strip().upper()
    codigo = codigo.replace("$", "S")
    codigo = codigo.replace(" ", "")
    codigo = re.sub(r"[^A-Z0-9\-]", "", codigo)
    return codigo


def codigo_valido(codigo):
    codigo = limpar_codigo(codigo)

    if len(codigo) < 4:
        return False

    if not re.search(r"\d", codigo):
        return False

    if codigo in [
        "0000", "00000", "20000",
        "0001", "0002", "0003", "0004", "0005",
        "0006", "0007", "0008", "0009", "0010",
        "2026", "18775", "005922", "8945",
        "2760", "9926"
    ]:
        return False

    if codigo.startswith("-"):
        return False

    if codigo.endswith("-"):
        return False

    if codigo.isdigit() and len(codigo) < 5:
        return False

    return True


def extrair_codigos_do_texto(texto):
    texto = str(texto).upper()

    encontrados = re.findall(
        r"\b[A-Z]{1,4}\d{5,15}[A-Z0-9]*\b",
        texto
    )

    codigos = []

    for codigo in encontrados:
        codigo = limpar_codigo(codigo)

        if codigo_valido(codigo):
            codigos.append(codigo)

    return codigos

    return True
def extrair_codigos_do_texto(texto):
    texto = str(texto).upper()

    encontrados = re.findall(
        r"\b[A-Z]{1,4}\d{5,15}[A-Z0-9]*\b",
        texto
    )

    codigos = []

    for codigo in encontrados:
        codigo = limpar_codigo(codigo)

        if codigo_valido(codigo):
            codigos.append(codigo)

    return codigos

def limpar_itens(itens):
    itens_limpos = []

    for item in itens:
        codigo = str(item.get("codigo", "")).strip().upper()
        qtd = str(item.get("qtd", "1")).strip()
        descricao_original = str(item.get("descricao", "")).strip().upper()

        codigo = codigo.replace(",", " / ")
        codigo = codigo.replace(";", " / ")
        codigo = codigo.replace("|", " / ")

        partes_codigo = re.split(r"\s*/\s*|\s+", codigo)

        codigos_validos = []

        for parte in partes_codigo:
            parte = limpar_codigo(parte)

            if codigo_valido(parte):
                codigos_validos.append(parte)

        codigos_da_descricao = extrair_codigos_do_texto(descricao_original)

        for codigo_desc in codigos_da_descricao:
            if codigo_desc not in codigos_validos:
                codigos_validos.append(codigo_desc)

        if not codigos_validos:
            continue

        try:
            qtd_num = float(qtd.replace(",", "."))
        except:
            qtd = "1"
            qtd_num = 1

        if qtd_num < 1 or qtd_num > 99:
            qtd = "1"

        codigo_final = " / ".join(codigos_validos)

        itens_limpos.append({
            "codigo": codigo_final,
            "qtd": qtd,
            "descricao": "ELEMENTO FILTRO"
        })

    return itens_limpos


def extrair_itens_com_visao(caminho_imagem):
    with open(caminho_imagem, "rb") as arquivo:
        imagem_base64 = base64.b64encode(arquivo.read()).decode("utf-8")

    extensao = os.path.splitext(caminho_imagem)[1].lower()

    mime = "image/jpeg"

    if extensao == ".png":
        mime = "image/png"

    if extensao == ".webp":
        mime = "image/webp"

    prompt = """
Analise esta imagem de cotacao, requisicao, lista de materiais ou planilha.

Extraia APENAS os itens reais da tabela.

REGRA MAIS IMPORTANTE:
- cada linha real da tabela deve virar apenas 1 item na resposta
- se uma mesma linha tiver mais de um codigo, coloque todos na mesma celula codigo, separados por " / "
- nao crie uma linha separada para cada codigo da mesma linha
- a descricao deve ser sempre "ELEMENTO FILTRO"

Ignore:
- cabecalho
- cliente
- endereco
- telefone
- data
- numero da requisicao
- solicitante
- rodape
- observacoes
- valores
- marcas isoladas
- titulos de coluna

Extraia somente:
- codigo
- qtd
- descricao

Regras:
- codigo pode ser numerico ou alfanumerico
- codigo nao e a sequencia do item 0001, 0002, 0003
- quantidade normalmente fica entre 1 e 99
- quantidade pode repetir varias vezes
- descricao sempre deve ser "ELEMENTO FILTRO"
- nao invente dados
- retorne somente JSON valido
- nao use markdown

Formato obrigatorio:
{
  "itens": [
    {
      "codigo": "CODIGO1 / CODIGO2",
      "qtd": "1",
      "descricao": "ELEMENTO FILTRO"
    }
  ]
}
"""

    resposta = client.chat.completions.create(
        model="google/gemini-2.0-flash-001",
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": prompt
                    },
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": "data:" + mime + ";base64," + imagem_base64
                        }
                    }
                ]
            }
        ],
        temperature=0,
        max_tokens=2000
    )

    conteudo = resposta.choices[0].message.content

    if not conteudo:
        return []

    conteudo = conteudo.strip()
    conteudo = conteudo.replace("```json", "")
    conteudo = conteudo.replace("```", "")
    conteudo = conteudo.strip()

    dados = json.loads(conteudo)

    return limpar_itens(dados.get("itens", []))


def preparar_imagem_para_ocr(caminho_imagem):
    imagem = Image.open(caminho_imagem)
    imagem = imagem.convert("L")
    imagem = ImageOps.autocontrast(imagem)
    imagem = ImageEnhance.Contrast(imagem).enhance(3)
    imagem = imagem.resize((imagem.width * 2, imagem.height * 2))
    imagem = imagem.filter(ImageFilter.SHARPEN)
    return imagem


def ler_texto_imagem(caminho_imagem):
    imagem = preparar_imagem_para_ocr(caminho_imagem)

    texto = pytesseract.image_to_string(
        imagem,
        lang="eng",
        config="--psm 6"
    )

    return texto


def ler_texto_pdf(caminho_pdf):
    texto_final = ""

    documento = fitz.open(caminho_pdf)

    for numero_pagina in range(len(documento)):
        pagina = documento[numero_pagina]
        texto_pagina = pagina.get_text()

        if texto_pagina and len(texto_pagina.strip()) > 30:
            texto_final += "\n" + texto_pagina
        else:
            pix = pagina.get_pixmap(matrix=fitz.Matrix(2, 2))

            nome_imagem = os.path.join(
                UPLOAD_FOLDER,
                "pagina_pdf_" + str(numero_pagina) + ".png"
            )

            pix.save(nome_imagem)

            texto_ocr = ler_texto_imagem(nome_imagem)
            texto_final += "\n" + texto_ocr

    documento.close()

    return texto_final


def ler_texto_arquivo(caminho_arquivo):
    extensao = os.path.splitext(caminho_arquivo)[1].lower()

    if extensao == ".pdf":
        return ler_texto_pdf(caminho_arquivo)

    return ler_texto_imagem(caminho_arquivo)


def eh_item_sequencial(valor):
    valor = valor.strip()

    if not valor.isdigit():
        return False

    if len(valor) > 4:
        return False

    try:
        numero = int(valor)
    except:
        return False

    return numero >= 1 and numero <= 999


def eh_quantidade(valor):
    valor = valor.strip().replace(",", ".")

    try:
        numero = float(valor)
    except:
        return False

    return numero >= 1 and numero <= 99


def limpar_descricao(texto):
    return "ELEMENTO FILTRO"


def classificar_linha_cotacao(linha):
    linha = linha.strip().upper()

    if not linha:
        return None

    tokens = re.findall(
        r"\d+,\d+|[A-Z0-9\-]+",
        linha
    )

    if not tokens:
        return None

    codigos = []
    qtd = None

    for token in tokens:
        token_limpo = limpar_codigo(token)

        if token_limpo in [
            "ITEM", "QTD", "PRODUTO", "EQUIVALENCIA",
            "DESCRICAO", "DESCRICADO", "CODIGO"
        ]:
            continue

        if eh_item_sequencial(token_limpo):
            continue

        if qtd is None and eh_quantidade(token):
            qtd = token
            continue

        if codigo_valido(token_limpo):
            codigos.append(token_limpo)

    codigos_da_linha = extrair_codigos_do_texto(linha)

    for codigo_desc in codigos_da_linha:
        if codigo_desc not in codigos:
            codigos.append(codigo_desc)

    if not codigos:
        return None

    if qtd is None:
        qtd = "1"

    codigo_final = " / ".join(codigos)

    return {
        "codigo": codigo_final,
        "qtd": qtd,
        "descricao": "ELEMENTO FILTRO"
    }


def extrair_itens_do_texto(texto):
    itens = []
    codigos_usados = set()

    linhas = texto.splitlines()

    for linha in linhas:
        item = classificar_linha_cotacao(linha)

        if not item:
            continue

        codigo = item.get("codigo", "")

        if not codigo:
            continue

        if codigo in codigos_usados:
            continue

        codigos_usados.add(codigo)
        itens.append(item)

    return limpar_itens(itens)


def gerar_planilha_cotacao(itens):
    nome_arquivo = (
        "cotacao_" +
        datetime.now().strftime("%Y%m%d_%H%M%S") +
        ".xlsx"
    )

    caminho_saida = os.path.join(EXPORT_FOLDER, nome_arquivo)

    if not os.path.exists(MODELO_COTACAO):
        raise Exception("modelo_cotacao.xlsx nao encontrado")

    shutil.copy(MODELO_COTACAO, caminho_saida)

    wb = load_workbook(caminho_saida)
    ws = wb.active

    linha_inicio = 11
    coluna_codigo = 2
    coluna_qtd = 3
    coluna_descricao = 4

    areas_mescladas = list(ws.merged_cells.ranges)

    for area in areas_mescladas:
        min_col, min_row, max_col, max_row = area.bounds

        toca_area = not (
            max_row < linha_inicio or
            min_row > 80 or
            max_col < coluna_codigo or
            min_col > coluna_descricao
        )

        if toca_area:
            ws.unmerge_cells(str(area))

    for index, item in enumerate(itens):
        linha = linha_inicio + index

        ws.cell(row=linha, column=coluna_codigo).value = item.get("codigo", "")
        ws.cell(row=linha, column=coluna_qtd).value = item.get("qtd", "")
        ws.cell(row=linha, column=coluna_descricao).value = "ELEMENTO FILTRO"

    wb.save(caminho_saida)

    return nome_arquivo


@app.route("/", methods=["GET", "POST"])
def home():
    if not usuario_logado():
        return redirect("/login")

    nome = session["usuario_nome"]
    resposta_agente = None
    historico = buscar_historico(nome)

    if request.method == "POST":
        mensagem = request.form["mensagem"]

        resposta_agente = responder_chat(
            mensagem,
            historico
        )

        salvar_mensagem(
            nome,
            mensagem,
            resposta_agente
        )

        historico = buscar_historico(nome)

    return render_template(
        "index.html",
        nome=nome,
        resposta_agente=resposta_agente,
        historico=historico
    )


@app.route("/cotacao", methods=["GET", "POST"])
def cotacao():
    if not usuario_logado():
        return redirect("/login")

    erro = None
    arquivo = None
    texto_ocr = None

    if request.method == "POST":
        print("######### TESTE BYTE MASTER #########", flush=True)

        arquivo_enviado = request.files.get("imagem")

        if not arquivo_enviado:
            erro = "Nenhum arquivo enviado"
        else:
            nome_seguro = secure_filename(arquivo_enviado.filename)

            caminho_arquivo = os.path.join(
                UPLOAD_FOLDER,
                nome_seguro
            )

            arquivo_enviado.save(caminho_arquivo)

            try:
                extensao = os.path.splitext(caminho_arquivo)[1].lower()

                itens = []

                if extensao in [".jpg", ".jpeg", ".png", ".webp"]:
                    try:
                        print("Tentando IA de visao...", flush=True)
                        itens = extrair_itens_com_visao(caminho_arquivo)
                        print("Itens pela visao:", itens, flush=True)

                    except Exception as erro_visao:
                        print("ERRO NA VISAO:", erro_visao, flush=True)
                        erro = "Erro na IA de visao: " + str(erro_visao)
                        itens = []

                texto_ocr = ""

                if os.name == "nt":
                    try:
                        texto_ocr = ler_texto_arquivo(caminho_arquivo)

                        if not itens:
                            itens = extrair_itens_do_texto(texto_ocr)

                    except Exception as erro_ocr:
                        print("OCR ignorado:", erro_ocr, flush=True)

                print("=" * 50, flush=True)
                print("TEXTO OCR:", flush=True)
                print(texto_ocr, flush=True)
                print("=" * 50, flush=True)
                print("ITENS ENCONTRADOS:", flush=True)
                print(itens, flush=True)
                print("=" * 50, flush=True)

                if not itens and erro is None:
                    erro = "Nao consegui encontrar itens"

                if itens:
                    arquivo = gerar_planilha_cotacao(itens)
                    erro = None

            except Exception as e:
                erro = str(e)
                print("ERRO GERAL:", e, flush=True)

    return render_template(
        "cotacao.html",
        erro=erro,
        arquivo=arquivo,
        texto_ocr=texto_ocr,
        nome=session.get("usuario_nome")
    )


@app.route("/download/<nome_arquivo>")
def download(nome_arquivo):
    caminho = os.path.join(EXPORT_FOLDER, nome_arquivo)

    return send_file(
        caminho,
        as_attachment=True
    )


@app.route("/login", methods=["GET", "POST"])
def login():
    erro = None

    if request.method == "POST":
        nome = request.form["nome"]
        senha = request.form["senha"]

        usuario = buscar_usuario(nome)

        if usuario and check_password_hash(
            usuario["senha"],
            senha
        ):
            session["usuario_nome"] = nome
            return redirect("/")

        erro = "Nome ou senha incorretos"

    return render_template(
        "login.html",
        erro=erro
    )


@app.route("/register", methods=["GET", "POST"])
def register():
    erro = None

    if request.method == "POST":
        nome = request.form["nome"]
        senha = request.form["senha"]

        if buscar_usuario(nome):
            erro = "Esse nome ja existe"
        else:
            criar_usuario(nome, senha)
            session["usuario_nome"] = nome
            return redirect("/")

    return render_template(
        "register.html",
        erro=erro
    )


@app.route("/logout")
def logout():
    session.clear()

    return redirect("/login")


if __name__ == "__main__":
    port = int(
        os.environ.get("PORT", 5000)
    )

    app.run(
        host="0.0.0.0",
        port=port
    )